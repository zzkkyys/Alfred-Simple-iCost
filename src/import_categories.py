#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
iCost Alfred Workflow - 从 Excel 导入分类
从 iCost 导出的账单 Excel 文件中读取分类信息
"""

import json
import sys
import os

# 添加 workflow 包路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from workflow import Workflow3

DATA_FILENAME = "icost_data.json"


def load_data(wf):
    """加载现有数据（从 cache 目录）"""
    data_file = wf.cachefile(DATA_FILENAME)
    if os.path.exists(data_file):
        with open(data_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {
        "accounts": [],
        "expense_categories": {},
        "income_categories": {}
    }


def save_data(wf, data):
    """保存数据（到 cache 目录）"""
    data_file = wf.cachefile(DATA_FILENAME)
    with open(data_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def import_from_excel(wf, file_path):
    """从 Excel 文件导入分类"""
    try:
        import openpyxl
    except ImportError:
        return None, "请先安装 openpyxl: pip3 install openpyxl"
    
    if not os.path.exists(file_path):
        return None, f"文件不存在: {file_path}"
    
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet = wb.active
        
        # 读取表头，找到对应列
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        
        # 查找列索引
        type_col = None
        cat1_col = None
        cat2_col = None
        account_col = None
        
        for idx, header in enumerate(headers):
            if header:
                header_lower = str(header).lower()
                if '类型' in header_lower or 'type' in header_lower:
                    type_col = idx
                elif '一级分类' in header_lower or '分类' == header_lower:
                    cat1_col = idx
                elif '二级分类' in header_lower or '子分类' in header_lower:
                    cat2_col = idx
                elif '账户' in header_lower or 'account' in header_lower:
                    account_col = idx
        
        if cat1_col is None:
            # 尝试更宽松的匹配
            for idx, header in enumerate(headers):
                if header and '分类' in str(header):
                    if cat1_col is None:
                        cat1_col = idx
                    elif cat2_col is None:
                        cat2_col = idx
        
        # 读取数据
        expense_categories = {}
        income_categories = {}
        accounts = set()
        
        for row in sheet.iter_rows(min_row=2):
            # 获取类型（支出/收入）
            record_type = ""
            if type_col is not None and row[type_col].value:
                record_type = str(row[type_col].value).strip()
            
            # 获取一级分类
            cat1 = ""
            if cat1_col is not None and row[cat1_col].value:
                cat1 = str(row[cat1_col].value).strip()
            
            # 获取二级分类
            cat2 = ""
            if cat2_col is not None and row[cat2_col].value:
                cat2 = str(row[cat2_col].value).strip()
            
            # 获取账户
            if account_col is not None and row[account_col].value:
                accounts.add(str(row[account_col].value).strip())
            
            if not cat1:
                continue
            
            # 根据类型分类
            if '收入' in record_type or 'income' in record_type.lower():
                if cat1 not in income_categories:
                    income_categories[cat1] = []
                if cat2 and cat2 not in income_categories[cat1]:
                    income_categories[cat1].append(cat2)
            else:
                # 默认为支出
                if cat1 not in expense_categories:
                    expense_categories[cat1] = []
                if cat2 and cat2 not in expense_categories[cat1]:
                    expense_categories[cat1].append(cat2)
        
        wb.close()
        
        # 合并现有数据
        existing_data = load_data(wf)
        
        # 合并账户
        all_accounts = list(set(list(existing_data.get("accounts", [])) + list(accounts)))
        
        # 合并分类
        for cat1, cat2_list in expense_categories.items():
            if cat1 in existing_data.get("expense_categories", {}):
                existing_data["expense_categories"][cat1] = list(set(
                    existing_data["expense_categories"][cat1] + cat2_list
                ))
            else:
                if "expense_categories" not in existing_data:
                    existing_data["expense_categories"] = {}
                existing_data["expense_categories"][cat1] = cat2_list
        
        for cat1, cat2_list in income_categories.items():
            if cat1 in existing_data.get("income_categories", {}):
                existing_data["income_categories"][cat1] = list(set(
                    existing_data["income_categories"][cat1] + cat2_list
                ))
            else:
                if "income_categories" not in existing_data:
                    existing_data["income_categories"] = {}
                existing_data["income_categories"][cat1] = cat2_list
        
        existing_data["accounts"] = all_accounts
        
        # 保存数据到 cache 目录
        save_data(wf, existing_data)
        
        expense_count = sum(len(v) for v in expense_categories.values())
        income_count = sum(len(v) for v in income_categories.values())
        
        return existing_data, f"导入成功！支出分类: {len(expense_categories)} 个一级分类，{expense_count} 个二级分类；收入分类: {len(income_categories)} 个一级分类，{income_count} 个二级分类；账户: {len(accounts)} 个"
        
    except Exception as e:
        return None, f"导入失败: {str(e)}"

def main(wf):
    # 获取用户输入的文件路径
    query = wf.args[0].strip() if wf.args else ""
    
    if query:
        # 用户提供了文件路径
        file_path = query
        
        # 处理路径中的 ~ 和空格
        file_path = os.path.expanduser(file_path)
        
        if os.path.exists(file_path):
            wf.add_item(
                title=f"📥 导入分类: {os.path.basename(file_path)}",
                subtitle=f"从 Excel 文件导入分类数据",
                arg=file_path,
                uid="import",
                icon="icon.png",
                valid=True
            )
        else:
            wf.add_item(
                title="⚠️ 文件不存在",
                subtitle=f"请检查路径: {file_path}",
                uid="not_found",
                icon="icon.png",
                valid=False
            )
    else:
        # 显示使用说明
        wf.add_item(
            title="📥 导入 iCost 分类",
            subtitle="请输入 Excel 文件路径，或将文件拖拽到这里",
            uid="help",
            icon="icon.png",
            valid=False
        )
        
        wf.add_item(
            title="💡 提示",
            subtitle="Excel 文件需包含: 类型、一级分类、二级分类 列",
            uid="tip",
            icon="icon.png",
            valid=False
        )
    
    wf.send_feedback()


if __name__ == "__main__":
    wf = Workflow3()
    sys.exit(wf.run(main))
